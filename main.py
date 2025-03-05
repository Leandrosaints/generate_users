import sys
import csv
import json
import re
import subprocess
import time

import pandas as pd
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFrame, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QDialog, QListWidget, QRadioButton, QComboBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QPixmap, QFont, QIcon, QMovie
from openpyxl.reader.excel import load_workbook
from text_unidecode import unidecode
import requests

# Executa o comando Powershell
def executar_comando(comando):
    try:
        result = subprocess.run(
            comando,
            capture_output=True, text=True, shell=True
        )
        return result.stdout, result.stderr
    except Exception as e:
        return "", str(e)

class OUSelectionDialog(QDialog):
    """Janela de diálogo para selecionar uma OU."""
    def __init__(self, ous, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Selecionar OU")
        self.setFixedSize(400, 300)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Lista de OUs
        self.ou_list = QListWidget(self)
        self.ou_list.addItems(ous)
        self.layout.addWidget(self.ou_list)

        # Botão de confirmação
        self.btn_confirmar = QPushButton("Selecionar", self)
        self.btn_confirmar.clicked.connect(self.accept)
        self.layout.addWidget(self.btn_confirmar)

    def get_selected_ou(self):
        """Retorna a OU selecionada."""
        return self.ou_list.currentItem().text()
class LoadingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Carregando...")
        self.setModal(True)  # Torna o diálogo modal
        self.setFixedSize(300, 100)  # Tamanho fixo do diálogo

        # Layout horizontal
        layout = QHBoxLayout()

        # Label para o texto
        self.label = QLabel("Por favor, aguarde...", self)
        #self.label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # Alinhamento do texto

        # Label para o GIF de loading
        self.loading_gif = QLabel(self)
        self.movie = QMovie("src/gif_load.gif")  # Substitua pelo caminho do seu GIF
        self.loading_gif.setMovie(self.movie)
        self.movie.start()  # Inicia o GIF

        # Adiciona os widgets ao layout
        layout.addWidget(self.loading_gif)
        layout.addWidget(self.label)

        self.setLayout(layout)


class ExcelProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        font = QFont('Arial', 12)
        self.setWindowTitle("GenServ")
        self.setWindowIcon(QIcon("src/ico.ico"))
        self.setStyleSheet("background-color: #F0F0F0;")  # Fundo da janela
        self.loading_dialog = LoadingDialog(self)
        # Atributo para armazenar o DataFrame final
        self.df_final = None

        # Layout principal
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QHBoxLayout()
        self.central_widget.setLayout(self.layout)

        # Barra lateral (frame esquerdo)
        self.sidebar_frame = QFrame()
        self.sidebar_frame.setStyleSheet("""
            background-color: #2E8BC0;
            border-radius: 20px;  /* Bordas mais arredondadas */
            padding: 20px;        /* Mais espaço interno */
            margin: 10px;         /* Margem externa */
            border: 2px solid #1C6EA4;  /* Borda sutil */
        """)
        self.sidebar_layout = QVBoxLayout()
        self.sidebar_frame.setLayout(self.sidebar_layout)

        # Adicionando botões à barra lateral
        self.btn_fetch_users = QPushButton("Buscar Usuários", self)
        self.btn_mover_usuarios = QPushButton("Mover Usuários", self)
        self.btn_excluir_usuarios = QPushButton("Excluir Usuários", self)
        self.btn_process = QPushButton("Processar", self)
        self.btn_download = QPushButton("Baixar Planilha", self)
        self.btn_powershell_run = QPushButton("Executar PowerShell", self)
        self.btn_add_ps1 = QPushButton("Adicionar Script .ps1", self)
        self.usuarios_selecionados = []
        self.dict_remove = {}

        # Adiciona os botões à barra lateral
        for button in [
            self.btn_process, self.btn_fetch_users, self.btn_mover_usuarios,
            self.btn_excluir_usuarios, self.btn_powershell_run, self.btn_add_ps1,
            self.btn_download,
        ]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #FFFFFF;
                    color: #2E8BC0;
                    border: 2px solid #2E8BC0;
                    border-radius: 5px;
                    padding: 10px;
                    font-size: 14px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2E8BC0;
                    color: #FFFFFF;
                }
            """)
            self.sidebar_layout.addWidget(button)

        # Adiciona a barra lateral ao layout principal
        self.layout.addWidget(self.sidebar_frame)

        # Frame direito (conteúdo principal)
        self.right_frame = QFrame()
        self.right_layout = QVBoxLayout()
        self.right_frame.setLayout(self.right_layout)

        # Título
        self.title_label = QLabel("GenServ - User Generation", self)
        self.title_label.setStyleSheet("font-size: 24px; font-weight: bold; color: #2E8BC0;")
        self.right_layout.addWidget(self.title_label)

        # Label de loading
        self.loading_label = QLabel(self)
        self.loading_movie = QMovie("src/gif_load.gif")
        self.loading_label.setMovie(self.loading_movie)
        self.loading_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        self.right_layout.addWidget(self.loading_label, alignment=Qt.AlignmentFlag.AlignRight)

        # Campos de entrada
        self.file_label = QLabel('Arquivo Excel:')
        self.file_input = QLineEdit(self)
        self.file_button = QPushButton('Escolher Arquivo')
        self.file_button.setStyleSheet("""
            QPushButton {
                background-color: #2E8BC0;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #2A6B94;
            }
        """)
        self.file_button.clicked.connect(self.open_file_dialog)

        self.domain_label = QLabel('Domínio:')
        self.domain_input = QLineEdit('@alunosenai.mt')
        self.office_label = QLabel('Office:')
        self.office_input = QLineEdit('SENAI - Nova Mutum/MT')
        self.creator_label = QLabel('Criador:')
        self.creator_input = QLineEdit("Criado Nome TI")
        self.dest_label = QLabel('Destino:')
        self.dest_input = QLineEdit('OU=CURSOS,OU=SENAINMT,OU=SENAI,DC=SFIEMT-EDU,DC=,DC=')

        # Contador de usuários gerados
        self.counter_label = QLabel('USER GERADOS: 0', self)
        self.counter_label.setStyleSheet("font-weight: bold; color: #2E8BC0;")

        # Campo de busca por CPF
        self.busca_cpf_input = QLineEdit(self)
        self.busca_cpf_input.setPlaceholderText("Digite o CPF para buscar...")
        self.busca_cpf_input.textChanged.connect(self.filtrar_por_cpf)

        # Opções de busca (Tabela ou AD)
        self.opcao_busca_tabela_radio = QRadioButton("Buscar na Tabela", self)
        self.opcao_busca_ad_radio = QRadioButton("Buscar no AD", self)
        self.opcao_busca_tabela_radio.setChecked(True)  # Tabela selecionada por padrão

        # Combo de filtros para busca no AD
        self.filtro_ad_combo = QComboBox(self)
        self.filtro_ad_combo.addItems(["CPF", "OU", "Nome", "E-mail", "Departamento", "Status"])
        self.filtro_ad_combo.setVisible(False)  # Inicialmente oculto

        # Layout para opções de busca
        busca_opcoes_layout = QHBoxLayout()
        busca_opcoes_layout.addWidget(self.opcao_busca_tabela_radio)
        busca_opcoes_layout.addWidget(self.opcao_busca_ad_radio)
        busca_opcoes_layout.addWidget(self.filtro_ad_combo)

        # Adiciona widgets ao layout direito
        self.right_layout.addWidget(self.file_label)
        self.right_layout.addWidget(self.file_input)
        self.right_layout.addWidget(self.file_button)
        self.right_layout.addWidget(self.domain_label)
        self.right_layout.addWidget(self.domain_input)
        self.right_layout.addWidget(self.office_label)
        self.right_layout.addWidget(self.office_input)
        self.right_layout.addWidget(self.creator_label)
        self.right_layout.addWidget(self.creator_input)
        self.right_layout.addWidget(self.dest_label)
        self.right_layout.addWidget(self.dest_input)
        self.right_layout.addWidget(self.counter_label)
        self.right_layout.addWidget(self.busca_cpf_input)
        self.right_layout.addLayout(busca_opcoes_layout)

        # Tabela para exibir usuários existentes no AD
        self.table_existing_users = QTableWidget(self)
        self.table_existing_users.setColumnCount(3)
        self.table_existing_users.setHorizontalHeaderLabels(["Nome", "CPF", "Localização"])
        self.table_existing_users.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.right_layout.addWidget(self.table_existing_users)

        # Adiciona o frame direito ao layout principal
        self.layout.addWidget(self.right_frame)

        # Conecta os botões às funções
        self.btn_fetch_users.clicked.connect(self.fetch_existing_users)
        self.btn_mover_usuarios.clicked.connect(self.mover_usuarios)
        self.btn_excluir_usuarios.clicked.connect(self.excluir_usuarios)
        self.btn_process.clicked.connect(self.process_file)
        self.btn_download.clicked.connect(self.download_file)
        self.btn_powershell_run.clicked.connect(self.run_powershell)
        self.btn_add_ps1.clicked.connect(self.add_powershell_script)

        # Conecta a mudança de opção (Tabela ou AD) para mostrar/ocultar o combo de filtros
        self.opcao_busca_ad_radio.toggled.connect(self.filtro_ad_combo.setVisible)

    def show_loading_t(self):
        self.loading_dialog.show()  # Exibe o diálogo como modal

    def hide_loading_T(self):
        self.loading_dialog.hide()  # Oculta o diálogo

    def show_loading(self):
        """Inicia a animação de loading."""
        self.loading_movie.start()

    def hide_loading(self):
        """Para a animação de loading."""
        self.loading_movie.stop()
        self.loading_label.hide()  # Esconde o label de loading

    def download_file(self):
        self.show_loading()
        # URL do arquivo no Google Drive
        file_url = 'https://docs.google.com/spreadsheets/d/173_jF1J7_84oLtI_W0lIf30Yd8CGYaux/export?format=xlsx'
        response = requests.get(file_url)

        if response.status_code == 200:
            # Abre uma caixa de diálogo para o usuário escolher onde salvar o arquivo
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getSaveFileName(
                self,
                'Salvar Arquivo',  # Título da janela
                '',  # Diretório inicial (vazio para o diretório padrão)
                'Arquivos Excel (*.xlsx)'  # Filtro de extensão
            )

            # Se o usuário escolher um local e nome de arquivo
            if file_path:
                # Adiciona a extensão .xlsx se o usuário não a incluir
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'

                # Salva o arquivo no local escolhido
                with open(file_path, 'wb') as f:
                    f.write(response.content)
                QMessageBox.information(self, 'Sucesso', f'Arquivo salvo com sucesso em: {file_path}')
            else:
                QMessageBox.warning(self, 'Aviso', 'Nenhum local foi selecionado para salvar o arquivo.')
        else:
            QMessageBox.critical(self, 'Erro', 'Erro ao baixar o arquivo!')
        self.hide_loading()

    def add_powershell_script(self):
        self.show_loading()
        file_dialog = QFileDialog()
        ps_file_path, _ = file_dialog.getOpenFileName(self, 'Escolher Arquivo PowerShell', '',
                                                      'PowerShell Files (*.ps1)')

        if ps_file_path:
            try:
                comando = f'powershell -ExecutionPolicy Bypass -File "{ps_file_path}"'
                stdout, stderr = executar_comando(comando)

                if stderr:
                    QMessageBox.critical(self, 'Erro', f'Erro ao executar o PowerShell: {stderr}')
                else:
                    QMessageBox.information(self, 'Sucesso', 'Script PowerShell executado com sucesso.')
            except Exception as e:
                QMessageBox.critical(self, 'Erro', f'Erro ao executar o script PowerShell: {e}')
        self.hide_loading()

    def open_file_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, 'Escolher Arquivo', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            self.file_input.setText(file_path)

    def nome_sem_acento(self, nome):
        return unidecode(nome)

    def validar_entradas(self):
        input_file = self.file_input.text()
        if not input_file.endswith(('.xlsx', '.xls')):
            QMessageBox.warning(self, 'Erro', 'Por favor, selecione um arquivo Excel válido (.xlsx ou .xls).')
            return False

        dominio = self.domain_input.text()
        if '@' not in dominio:
            QMessageBox.warning(self, 'Erro', 'Por favor, insira um domínio válido (ex: @alunosenai.mt).')
            return False

        office = self.office_input.text()
        if not office.strip():
            QMessageBox.warning(self, 'Erro', 'O campo "Office" não pode estar vazio.')
            return False

        criador = self.creator_input.text()
        if not criador.strip():
            QMessageBox.warning(self, 'Erro', 'O campo "Criador" não pode estar vazio.')
            return False

        destino = self.dest_input.text()
        if not destino.strip():
            QMessageBox.warning(self, 'Erro', 'O campo "Destino" não pode estar vazio.')
            return False

        return True

    def process_file(self):
        self.show_loading_t()
        input_file = self.file_input.text()
        dominio = self.domain_input.text()
        office = self.office_input.text()
        criador = self.creator_input.text()
        destino = self.dest_input.text()

        if not self.validar_entradas():
            self.hide_loading_T()
            return

        try:
            df = pd.read_excel(input_file)
            df.columns = ['TURMA', 'RA', 'ALUNO', 'CPF', 'USUÁRIO', 'SENHA', 'E-MAIL / OFFICE 365']
            df = df[~df['RA'].astype(str).str.contains('RA', case=False)]
            df = df.dropna(subset=['RA', 'ALUNO'])

            # Armazenar o DataFrame final como atributo da classe
            self.df_final = pd.DataFrame()

            # Processar os dados...
            nome_count = df['ALUNO'].count()  # Contando nomes na coluna 'ALUNO'
            self.counter_label.setText(f'USER GERADOS: {nome_count}')  # Atualiza o contador

            def gerar_senha(ra):
                return "00" + ra[2:]

            def separar_nome(nome_completo):
                nome_completo_sem_acento = self.nome_sem_acento(nome_completo)
                nome_limpo = re.sub(r'[^\w\s]', '', nome_completo_sem_acento)
                partes = nome_limpo.split()
                primeiro_nome = partes[0].title() if partes else ''
                sobrenome = ' '.join(part.title() for part in partes[1:]) if len(partes) > 1 else ''
                return primeiro_nome, sobrenome

            def formatar_nome(nome):
                nome_sem_acentos = self.nome_sem_acento(nome)
                nome_formatado = nome_sem_acentos.title()
                return nome_formatado

            df['Senha'] = df['RA'].apply(lambda x: gerar_senha(str(x)))
            df['Primeiro Nome'], df['Sobrenome'] = zip(*df['ALUNO'].apply(separar_nome))
            df['E-mail'] = df['RA'].apply(lambda x: str(x).zfill(8) + dominio)
            df['Descrição Completa'] = df['ALUNO']

            # Criar DataFrame final
            self.df_final['Nome'] = df['ALUNO'].apply(formatar_nome)
            self.df_final['Dn'] = df['Descrição Completa'].apply(formatar_nome) + ' - ' + office
            self.df_final['PrimeiroNome'] = df['Primeiro Nome']
            self.df_final['Sobrenome'] = df['Sobrenome']
            self.df_final['Conta'] = df['RA'].apply(lambda x: str(x).zfill(8))
            self.df_final['Email'] = df['E-mail']
            self.df_final['Desc'] = df['CPF']
            self.df_final['Office'] = office
            self.df_final['Dep'] = criador
            self.df_final['OU'] = destino
            self.df_final['Pass'] = df['Senha']

            # Abre uma caixa de diálogo para o usuário escolher onde salvar o arquivo
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getSaveFileName(
                self,
                'Salvar Arquivo CSV',  # Título da janela
                '',  # Diretório inicial (vazio para o diretório padrão)
                'Arquivos CSV (*.csv)'  # Filtro de extensão
            )

            # Se o usuário escolher um local e nome de arquivo
            if file_path:
                # Adiciona a extensão .csv se o usuário não a incluir
                if not file_path.endswith('.csv'):
                    file_path += '.csv'

                # Salva o arquivo no local escolhido
                with open(file_path, 'w', encoding='utf-8') as file:
                    header = 'Nome,Dn,PrimeiroNome,Sobrenome,Conta,Email,Desc,Office,Dep,OU,Pass\n'
                    file.write(header)
                    for _, row in self.df_final.iterrows():
                        line = (
                            f'"{row["Nome"]}",'
                            f'"{row["Dn"]}",'
                            f'{row["PrimeiroNome"]},'
                            f'{row["Sobrenome"]},'
                            f'{row["Conta"]},'
                            f'{row["Email"]},'
                            f'{row["Desc"]},'
                            f'"{row["Office"]}",'
                            f'"{row["Dep"]}",'
                            f'"{row["OU"]}",'
                            f'"{row["Pass"]}"\n'
                        )
                        file.write(line)

                self.adicionar_dados_planilha(input_file)
                QMessageBox.information(self, 'Sucesso', f'Arquivo salvo com sucesso em: {file_path}')
                # Pergunta ao usuário se deseja buscar usuários duplicados
                resposta = QMessageBox.question(
                    self,
                    'Buscar Usuários Duplicados',
                    'Deseja fazer uma busca de usuários duplicados?',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if resposta == QMessageBox.StandardButton.Yes:
                    # Executa a função para buscar usuários duplicados
                    self.fetch_existing_users()
                else:
                    QMessageBox.information(self, 'OK', 'Operação concluída.')
            else:
                QMessageBox.warning(self, 'Aviso', 'Nenhum local foi selecionado para salvar o arquivo.')

        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao processar o arquivo: {e}, verifique a estrutura da planilha!')
        self.hide_loading_T()

    def adicionar_dados_planilha(self, arquivo_excel):
        wb = load_workbook(arquivo_excel)
        sheet = wb.active
        proxima_linha = 3

        for conta, usuario, senha in zip(self.df_final["Conta"], self.df_final["Pass"], self.df_final["Email"]):
            sheet.cell(row=proxima_linha, column=5, value=conta)
            sheet.cell(row=proxima_linha, column=6, value=usuario)
            sheet.cell(row=proxima_linha, column=7, value=senha)
            proxima_linha += 1

        wb.save(arquivo_excel)

    def run_powershell(self):
        self.show_loading_t()
        try:
            # Abre uma caixa de diálogo para o usuário escolher o arquivo CSV
            file_dialog = QFileDialog()
            csv_file_path, _ = file_dialog.getOpenFileName(
                self,
                'Selecionar Arquivo CSV',  # Título da janela
                '',  # Diretório inicial (vazio para o diretório padrão)
                'Arquivos CSV (*.csv)'  # Filtro de extensão
            )

            # Se o usuário selecionar um arquivo
            if csv_file_path:
                # Monta o comando PowerShell para usar o arquivo selecionado
                comando = (
                    f'powershell -ExecutionPolicy Bypass '
                    f'-Command "Import-Csv \'{csv_file_path}\' | '
                    'ForEach-Object {New-ADUser -Name $_.nome -DisplayName $_.dn -givenName $_.primeironome '
                    '-Surname $_.Sobrenome -SamAccountName $_.conta -UserPrincipalName $_.email -description $_.Desc '
                    '-Office $_.Office -Department $_.Dep -Path $_.ou '
                    '-AccountPassword (ConvertTo-SecureString -AsPlainText $_.pass -Force) -Enabled $true}"'
                )

                # Executa o comando PowerShell
                stdout, stderr = executar_comando(comando)

                if stderr:
                    QMessageBox.critical(self, 'Erro', f'Erro ao executar o PowerShell: {stderr}')
                else:
                    QMessageBox.information(self, 'Sucesso', 'Comando PowerShell executado com sucesso.')
            else:
                QMessageBox.warning(self, 'Aviso', 'Nenhum arquivo CSV foi selecionado.')
        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao executar o comando PowerShell: {e}')
        self.hide_loading_T()

    def filtrar_por_cpf(self):
        """Filtra a tabela ou busca no AD com base no CPF digitado."""
        cpf_busca = self.busca_cpf_input.text().strip()  # Obtém o texto digitado

        # Verifica onde o usuário deseja buscar (Tabela ou AD)
        if self.opcao_busca_tabela_radio.isChecked():  # Busca na tabela
            self.filtrar_na_tabela(cpf_busca)
        elif self.opcao_busca_ad_radio.isChecked():  # Busca no AD
            self.filtrar_no_ad(filtro='cpf', valor=cpf_busca)

    def filtrar_na_tabela(self, cpf_busca):
        """Filtra a tabela com base no CPF digitado."""
        for row in range(self.table_existing_users.rowCount()):
            cpf_tabela = self.table_existing_users.item(row, 1).text()  # Obtém o CPF da linha
            # Mostra ou oculta a linha com base no CPF digitado
            if cpf_busca.lower() in cpf_tabela.lower():
                self.table_existing_users.setRowHidden(row, False)  # Mostra a linha
            else:
                self.table_existing_users.setRowHidden(row, True)  # Oculta a linha

    def filtrar_no_ad(self, filtro, valor):
        """Busca usuários no AD com base no filtro e valor especificados."""
        usuarios = self.buscar_usuarios_ad(filtro=filtro, valor=valor)
        if usuarios:
            self.preencher_tabela_com_usuarios(usuarios)  # Preenche a tabela com os resultados
        else:
            QMessageBox.information(self, 'Informação', 'Nenhum usuário encontrado no AD.')

    def buscar_usuarios_ad(self, filtro, valor, propriedades=None):
        """
        Busca usuários no Active Directory com base em um filtro e valor especificados,
        e preenche a tabela com os resultados.
        """
        if propriedades is None:
            propriedades = ["Name", "Description", "DistinguishedName"]

        try:
            self.show_loading_t()

            # Define o filtro do PowerShell com base no tipo de busca
            if filtro == 'cpf':
                filtro_ps = f'Description -eq "{valor}"'
            elif filtro == 'ou':
                filtro_ps = f'DistinguishedName -like "*{valor}*"'
            elif filtro == 'nome':
                filtro_ps = f'Name -like "*{valor}*"'
            elif filtro == 'email':
                filtro_ps = f'UserPrincipalName -like "*{valor}*"'
            elif filtro == 'departamento':
                filtro_ps = f'Department -eq "{valor}"'
            elif filtro == 'status':
                filtro_ps = f'Enabled -eq {"$true" if valor else "$false"}'
            else:
                raise ValueError("Filtro inválido. Use 'cpf', 'ou', 'nome', 'email', 'departamento' ou 'status'.")

            # Define as propriedades como uma string separada por vírgulas
            propriedades_str = ",".join(propriedades)

            # Monta o comando PowerShell com ConvertTo-Json
            comando = (
                f'powershell -ExecutionPolicy Bypass -Command '
                f'"Get-ADUser -Filter \\"{filtro_ps}\\" -Properties {propriedades_str} | '
                f'Select-Object {propriedades_str} | ConvertTo-Json"'
            )

            # Executa o comando
            stdout, stderr = executar_comando(comando)

            # Verifica se houve erro no comando
            if stderr:
                QMessageBox.critical(self, 'Erro', f'Erro ao buscar usuários: {stderr}')
                return

            # Verifica se a saída está vazia
            if not stdout.strip():
                QMessageBox.information(self, 'Informação', 'Nenhum usuário encontrado.')
                return

            # Converte a saída do PowerShell (JSON) em uma lista de dicionários
            try:
                usuarios = json.loads(stdout)
                # Certifique-se de que 'usuarios' seja uma lista
                if isinstance(usuarios, dict):  # Se for um único usuário, coloque-o em uma lista
                    usuarios = [usuarios]
                self.preencher_tabela_com_usuarios(usuarios)
            except json.JSONDecodeError as e:
                QMessageBox.critical(self, 'Erro', f'Erro ao processar a saída do PowerShell: {e}')
                return

        except Exception as e:
            # Captura qualquer exceção e exibe uma mensagem de erro
            QMessageBox.critical(self, 'Erro', f'Erro ao buscar usuários: {e}')

        finally:
            self.hide_loading_T()

    def processar_saida_powershell(self, stdout):
        """
        Processa a saída do PowerShell (texto normal) e retorna uma lista de dicionários.
        """
        usuarios = []
        linhas = stdout.strip().split('\n')

        for linha in linhas:
            # Divide a linha em colunas com base em espaços em branco
            colunas = [coluna.strip() for coluna in linha.split() if coluna.strip()]
            if len(colunas) >= 3:  # Verifica se há pelo menos 3 colunas (Nome, CPF, Localização)
                usuario = {
                    "Name": colunas[0],
                    "Description": colunas[1],
                    "DistinguishedName": " ".join(colunas[2:])  # Junta o restante como DistinguishedName
                }
                usuarios.append(usuario)

        return usuarios

    def preencher_tabela_com_usuarios(self, usuarios):
        """Preenche a tabela com os usuários encontrados."""
        self.table_existing_users.setRowCount(0)  # Limpa a tabela

        # Configura as colunas da tabela
        self.table_existing_users.setColumnCount(3)
        self.table_existing_users.setHorizontalHeaderLabels(["Nome", "CPF (Descrição)", "Localização"])

        # Verifica se a lista de usuários está vazia
        if not usuarios:
            QMessageBox.warning(self, 'Aviso', 'Nenhum usuário encontrado para exibir.')
            return

        # Preenche a tabela com os dados
        for i, usuario in enumerate(usuarios):
            nome = usuario.get('Name', '')
            descricao = usuario.get('Description', '')
            distinguished_name = usuario.get('DistinguishedName', '')

            # Adiciona uma nova linha à tabela
            self.table_existing_users.insertRow(i)
            self.table_existing_users.setItem(i, 0, QTableWidgetItem(nome))
            self.table_existing_users.setItem(i, 1, QTableWidgetItem(descricao))
            self.table_existing_users.setItem(i, 2, QTableWidgetItem(distinguished_name))

        # Ajusta a largura das colunas
        self.table_existing_users.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        # Habilita a seleção de linhas
        self.table_existing_users.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_existing_users.setSelectionMode(QTableWidget.SelectionMode.MultiSelection)

    def fetch_existing_users(self):
        """Busca usuários existentes no AD, compara com o CSV e exibe duplicados na tabela."""

        try:
            self.show_loading_t()
            # Abre uma caixa de diálogo para o usuário escolher o arquivo CSV
            file_dialog = QFileDialog()
            csv_file_path, _ = file_dialog.getOpenFileName(
                self,
                'Selecionar Arquivo CSV',  # Título da janela
                '',  # Diretório inicial (vazio para o diretório padrão)
                'Arquivos CSV (*.csv)'  # Filtro de extensão
            )

            if not csv_file_path:
                QMessageBox.warning(self, 'Aviso', 'Nenhum arquivo CSV foi selecionado.')
                return

            # Lê o arquivo CSV
            df_csv = pd.read_csv(csv_file_path)

            # Verifica se a coluna 'Desc' (CPF) existe no CSV
            if 'Desc' not in df_csv.columns:
                QMessageBox.critical(self, 'Erro', 'A coluna "Desc" (CPF) não foi encontrada no arquivo CSV.')
                return

            # Extrai os CPFs da coluna 'Desc'
            cpfs_csv = df_csv['Desc'].astype(str).tolist()

            # Comando PowerShell para buscar usuários no AD com CPFs que estão no CSV
            comando = (
                    'powershell -ExecutionPolicy Bypass -Command "'
                    'Get-ADUser -Filter * -Properties Name, Description, DistinguishedName | '
                    'Where-Object { $_.Description -in @(' + ','.join(f'"{cpf}"' for cpf in cpfs_csv) + ') } | '
                                                                                                        'Select-Object Name, Description, DistinguishedName | '
                                                                                                        'ConvertTo-Json"'
            )

            # Executa o comando
            stdout, stderr = executar_comando(comando)

            if stderr:
                QMessageBox.critical(self, 'Erro', f'Erro ao buscar usuários: {stderr}')
            else:
                # Converte a saída do PowerShell (JSON) em uma lista de dicionários
                try:
                    usuarios = json.loads(stdout)
                except json.JSONDecodeError:
                    QMessageBox.critical(self, 'Erro', 'Erro ao processar a saída do PowerShell.')
                    return

                # Configura a tabela com as colunas necessárias
                self.table_existing_users.setColumnCount(3)
                self.table_existing_users.setHorizontalHeaderLabels(
                    ["Nome", "CPF (Descrição)", "Localização"]
                )
                self.table_existing_users.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

                # Preenche a tabela com os dados
                self.table_existing_users.setRowCount(len(usuarios))
                for i, usuario in enumerate(usuarios):
                    nome = usuario['Name']
                    descricao = usuario['Description']
                    distinguished_name = usuario['DistinguishedName']

                    # Extrai a parte da localização a partir de "OU=CURSOS"
                    #localizacao = distinguished_name.split("OU")[-1].strip(",")
                    #localizacao = "OU=CURSOS" + localizacao  # Adiciona "OU=CURSOS" de volta

                    self.table_existing_users.setItem(i, 0, QTableWidgetItem(nome))
                    self.table_existing_users.setItem(i, 1, QTableWidgetItem(descricao))
                    self.table_existing_users.setItem(i, 2, QTableWidgetItem(distinguished_name))

                # Habilita a seleção de linhas
                self.table_existing_users.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                self.table_existing_users.setSelectionMode(QTableWidget.SelectionMode.MultiSelection)

                # Botão para adicionar usuários selecionados à lista
                self.btn_adicionar_selecionados = QPushButton("Adicionar Selecionados à Lista", self)
                self.btn_adicionar_selecionados.clicked.connect(self.adicionar_selecionados_a_lista)
                self.right_layout.addWidget(self.btn_adicionar_selecionados)

        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao buscar usuários: {e}')
        finally:
            self.hide_loading_T()


    def listar_ous(self):
        """Lista todas as OUs disponíveis no AD."""
        self.show_loading_t()
        comando = (
            'powershell -ExecutionPolicy Bypass -Command "'
            'Get-ADOrganizationalUnit -Filter * | Select-Object -ExpandProperty DistinguishedName'
            '"'
        )
        stdout, stderr = executar_comando(comando)

        if stderr:
            QMessageBox.critical(self, 'Erro', f'Erro ao listar OUs: {stderr}')
            self.hide_loading_T()
            return []

        # Processa a saída para obter a lista de OUs
        ous = [ou.strip() for ou in stdout.splitlines() if ou.strip()]
        return ous

    def adicionar_selecionados_a_lista(self):
        """
        Adiciona os usuários selecionados da tabela a uma lista.

        :return: Lista de DistinguishedNames dos usuários selecionados.
        """
        try:
            selecionados = self.table_existing_users.selectedItems()
            if not selecionados:
                print('Nenhum usuário selecionado.')
                return []

            for row in set(item.row() for item in selecionados):
                distinguished_name = self.table_existing_users.item(row, 2).text()  # Pega o DistinguishedName
                self.usuarios_selecionados.append(distinguished_name)

            # Exibe a mensagem de confirmação
            QMessageBox.information(self, "Sucesso",
                                    f'{len(self.usuarios_selecionados)} usuário(s) adicionados à lista.')

            print(f'{len(self.usuarios_selecionados)} usuário(s) adicionados à lista.')
            return self.usuarios_selecionados

        except Exception as e:
            self.hide_loading_T()
            QMessageBox.warning(self, 'Aviso', f'Erro ao adicionar usuários à lista: {e}')
            return []

        finally:
            self.hide_loading_T()

    def mover_usuarios(self):
        """
        Move um ou mais usuários para uma nova OU no Active Directory.
        """
        if not self.usuarios_selecionados:
            QMessageBox.warning(self, 'Aviso', 'Nenhum usuário selecionado.')
            return

        # Obtém a lista de OUs
        ous = self.listar_ous()
        if not ous:
            QMessageBox.warning(self, 'Aviso', 'Nenhuma OU encontrada no AD.')
            return

        # Abre a janela de diálogo para selecionar a OU de destino
        dialog = OUSelectionDialog(ous, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            destino_ou = dialog.get_selected_ou()

            # Lista para armazenar usuários que não foram movidos
            usuarios_nao_movidos = []

            # Executa o comando PowerShell para cada usuário selecionado
            for usuario in self.usuarios_selecionados[:]:  # Usamos [:] para iterar sobre uma cópia da lista
                comando_powershell = f'Move-ADObject -Identity "{usuario}" -TargetPath "{destino_ou}"'

                try:
                    self.show_loading_t()
                    resultado = subprocess.run(
                        ["powershell.exe", "-ExecutionPolicy", "Bypass", "-NoProfile", "-Command", comando_powershell],
                        capture_output=True,
                        text=True
                    )

                    if resultado.returncode != 0:
                        print(f"Erro ao mover usuário {usuario}:", resultado.stderr)
                        usuarios_nao_movidos.append(usuario)  # Adiciona à lista de não movidos
                    else:
                        print(f"Usuário {usuario} movido com sucesso:", resultado.stdout)
                        self.remover_usuario_da_tabela(usuario)  # Remove o usuário da tabela

                except Exception as e:
                    print(f"Erro ao executar PowerShell para o usuário {usuario}:", str(e))
                    usuarios_nao_movidos.append(usuario)  # Adiciona à lista de não movidos

            # Atualiza a lista de selecionados com os usuários que não foram movidos
            self.usuarios_selecionados = usuarios_nao_movidos

            self.hide_loading_T()

            if not usuarios_nao_movidos:
                QMessageBox.information(self, 'Sucesso', 'Saida sucesso.')
            else:
                QMessageBox.warning(self, 'Aviso',
                                    f'Alguns usuários não foram movidos: {", ".join(usuarios_nao_movidos)}')

    def remover_usuario_da_tabela(self, usuario_dn):
        """
        Remove um usuário específico da tabela pelo DistinguishedName.
        """
        for row in range(self.table_existing_users.rowCount() - 1, -1, -1):  # Percorre de trás para frente
            item = self.table_existing_users.item(row, 2)  # Obtém a célula da coluna do DistinguishedName
            if item and item.text() == usuario_dn:
                self.table_existing_users.removeRow(row)  # Remove pelo índice
                return  # Sai da função após remover o usuário

    def verificar_usuarios_movidos(self, usuarios, destino_ou):
        """Verifica se os usuários foram movidos corretamente para a OU de destino."""
        for usuario in usuarios:
            try:
                # Comando PowerShell para verificar a OU atual do usuário
                comando_verificacao = f"""
                Import-Module ActiveDirectory;
                $usuario = Get-ADUser -Identity '{usuario}' -Properties DistinguishedName;
                $usuario.DistinguishedName
                """
                resultado = subprocess.run(
                    ["powershell", "-ExecutionPolicy", "Bypass", "-Command", comando_verificacao],
                    capture_output=True,
                    text=True,
                    shell=True
                )

                if resultado.stderr:
                    print(f"Erro ao verificar usuário {usuario}: {resultado.stderr}")
                    return False

                # Obtém o DistinguishedName atual do usuário
                dn_atual = resultado.stdout.strip()

                # Verifica se a OU de destino está no DistinguishedName atual
                if destino_ou not in dn_atual:
                    print(f"Usuário {usuario} não foi movido para a OU correta.")
                    print(f"DN atual: {dn_atual}")
                    print(f"OU de destino: {destino_ou}")
                    return False
            except Exception as e:
                print(f"Erro ao verificar usuário {usuario}: {e}")
                return False
        return True
    def exportar_duplicados(self):
        """Exporta os usuários duplicados para um arquivo CSV."""
        self.show_loading()
        try:
            # Abre uma caixa de diálogo para o usuário escolher onde salvar o arquivo
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getSaveFileName(
                self,
                'Salvar Arquivo CSV',  # Título da janela
                '',  # Diretório inicial (vazio para o diretório padrão)
                'Arquivos CSV (*.csv)'  # Filtro de extensão
            )

            # Se o usuário escolher um local e nome de arquivo
            if file_path:
                # Adiciona a extensão .csv se o usuário não a incluir
                if not file_path.endswith('.csv'):
                    file_path += '.csv'

                # Salva o arquivo no local escolhido
                with open(file_path, 'w', encoding='utf-8') as file:
                    header = 'Nome,CPF (Descrição),Localização\n'
                    file.write(header)
                    for i in range(self.table_existing_users.rowCount()):
                        nome = self.table_existing_users.item(i, 0).text()
                        cpf = self.table_existing_users.item(i, 1).text()
                        localizacao = self.table_existing_users.item(i, 2).text()
                        line = f'"{nome}","{cpf}","{localizacao}"\n'
                        file.write(line)

                QMessageBox.information(self, 'Sucesso', f'Arquivo salvo com sucesso em: {file_path}')
            else:
                QMessageBox.warning(self, 'Aviso', 'Nenhum local foi selecionado para salvar o arquivo.')
        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao exportar usuários duplicados: {e}')
        self.hide_loading()

    def excluir_usuarios(self):
        """Exclui usuários do AD e os remove da tabela."""
        self.show_loading()
        try:
            # Verifica se há usuários selecionados
            if not hasattr(self, 'usuarios_selecionados') or not self.usuarios_selecionados:
                QMessageBox.warning(self, 'Aviso', 'Nenhum usuário selecionado para excluir.')
                return

            # Lista para armazenar usuários que não foram excluídos
            usuarios_nao_excluidos = []

            # Executa o comando PowerShell para cada usuário selecionado
            for usuario in self.usuarios_selecionados[:]:  # Usamos [:] para iterar sobre uma cópia da lista
                comando_powershell = f'Remove-ADUser -Identity "{usuario}" -Confirm:$false'

                try:
                    resultado = subprocess.run(
                        ["powershell.exe", "-ExecutionPolicy", "Bypass", "-NoProfile", "-Command", comando_powershell],
                        capture_output=True,
                        text=True
                    )

                    if resultado.returncode != 0:
                        print(f"Erro ao excluir usuário {usuario}:", resultado.stderr)
                        usuarios_nao_excluidos.append(usuario)  # Adiciona à lista de não excluídos
                    else:
                        print(f"Usuário {usuario} excluído com sucesso:", resultado.stdout)
                        self.remover_usuario_da_tabela(usuario)  # Remove o usuário da tabela

                except Exception as e:
                    print(f"Erro ao executar PowerShell para o usuário {usuario}:", str(e))
                    usuarios_nao_excluidos.append(usuario)  # Adiciona à lista de não excluídos

            # Atualiza a lista de selecionados com os usuários que não foram excluídos
            self.usuarios_selecionados = usuarios_nao_excluidos

            if not usuarios_nao_excluidos:
                QMessageBox.information(self, 'Sucesso', 'excluído com sucesso.')
            else:
                QMessageBox.warning(self, 'Aviso',
                                    f'Alguns usuários não foram excluídos: {", ".join(usuarios_nao_excluidos)}')

        except Exception as e:
            QMessageBox.critical(self, 'Erro', f'Erro ao excluir usuários: {e}')

        finally:
            self.hide_loading()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    try:
        window = ExcelProcessor()
        window.showMaximized()
        sys.exit(app.exec())
    except Exception as e:
        print(f"Ocorreu um erro: {e}")